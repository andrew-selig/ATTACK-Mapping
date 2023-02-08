from openpyxl import load_workbook
import json
import argparse
import shutil

def process_xlsx(excel, template, output):
    create_output(template, output)
    wb = load_workbook(excel, data_only=True)
    ws = wb['Techniques']
    with open(output, 'r+') as attack_json:
        file_data = json.load(attack_json)
        for row in ws.iter_rows(min_row=2, max_col=6):
            technique_id = ''
            technique_score = ''
            for cell in row:
                if(cell.column_letter == "A"):
                    technique_id = cell.value
                elif(cell.column_letter == "F"):
                    technique_score = cell.value
            technique ={
                "techniqueID": technique_id,
                "score": technique_score
            }
            file_data['techniques'].append(technique)
            attack_json.seek(0)
            json.dump(file_data, attack_json, indent=6)
    
def create_output(template, output):
    shutil.copy(template, output)

parser = argparse.ArgumentParser()

parser.add_argument('--excel', '-e', type=str, default='examples/ATTACK-example.xlsx')
parser.add_argument('--template', '-t', type=str, default='examples/ATTACK-default.json')
parser.add_argument('--output', '-o', type=str, default='examples/ATTACK-output.json')

args = parser.parse_args()
process_xlsx(args.excel, args.template, args.output)